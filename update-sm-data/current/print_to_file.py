from openpyxl import load_workbook
from openpyxl.styles import numbers

def print_to_file(data, file_path):
	wb = load_workbook(file_path)
	
	daily_metrics_sheet = wb['Daily Metrics']
	
	daily_metrics_sheet.cell(row=11, column=2, value=data['yday_reach']).number_format = numbers.FORMAT_NUMBER
	daily_metrics_sheet.cell(row=12, column=2, value=f'={data["yday_views"]}-C12').number_format = numbers.FORMAT_NUMBER
	
	all_time_metrics_sheet = wb['All Time Metrics']
	
	all_time_metrics_sheet.cell(row=11, column=3, value=data['all_time_reach']).number_format = numbers.FORMAT_NUMBER
	all_time_metrics_sheet.cell(row=12, column=3, value=data['all_time_views']).number_format = numbers.FORMAT_NUMBER
	
	wb.save(file_path)
