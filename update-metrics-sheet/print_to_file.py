from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import date, timedelta
from typing import Any

def print_yesterday_to_file(wb: Workbook, yesterday: dict[str, Any]):
	sheet = wb['Daily Metrics']
	
	yesterdays_date = date.today() - timedelta(days=1)
	
	country_idx_manager = {}
	max_row = 15
	while sheet.cell(row=max_row, column=1).value:
		country_idx_manager[sheet.cell(row=max_row, column=1).value] = max_row
		max_row += 1
	max_col = sheet.max_column
	
	for col in range(max_col, 1, -1):
		row_1 = sheet.cell(row=1, column=col).value
		row_4 = sheet.cell(row=4, column=col).value
		row_5 = sheet.cell(row=5, column=col).value
		row_6 = sheet.cell(row=6, column=col).value
		row_7 = sheet.cell(row=7, column=col).value
		row_8 = sheet.cell(row=8, column=col).value
		row_9 = sheet.cell(row=9, column=col).value
		row_10 = sheet.cell(row=10, column=col).value
		row_11 = sheet.cell(row=11, column=col).value
		row_12 = sheet.cell(row=12, column=col).value
		if col != max_col:
			row_12 = str(row_12)
			row_12 = row_12[:row_12.index('-') + 1] + get_column_letter(column_index_from_string(row_12[row_12.index('-') + 1:len(row_12) - 2]) + 1) + '12'
		row_14 = sheet.cell(row=14, column=col).value
		
		sheet.cell(row=1, column=col + 1, value=row_1).number_format = numbers.FORMAT_DATE_YYYYMMDD2
		if col == max_col:
			sheet.cell(row=2, column=col + 1, value=f'=IFERROR({get_column_letter(col + 1)}9/{get_column_letter(col + 1)}11, "N/A")').number_format = numbers.FORMAT_PERCENTAGE
			sheet.cell(row=3, column=col + 1, value=f'=IFERROR({get_column_letter(col + 1)}7/{get_column_letter(col + 1)}12, "N/A")').number_format = numbers.FORMAT_PERCENTAGE
		sheet.cell(row=4, column=col + 1, value=row_4).number_format = numbers.FORMAT_PERCENTAGE
		sheet.cell(row=5, column=col + 1, value=row_5).number_format = numbers.FORMAT_PERCENTAGE
		sheet.cell(row=6, column=col + 1, value=row_6).number_format = numbers.FORMAT_NUMBER
		sheet.cell(row=7, column=col + 1, value=row_7).number_format = numbers.FORMAT_NUMBER
		sheet.cell(row=8, column=col + 1, value=row_8).number_format = numbers.FORMAT_NUMBER
		sheet.cell(row=9, column=col + 1, value=row_9).number_format = numbers.FORMAT_NUMBER
		sheet.cell(row=10, column=col + 1, value=row_10).number_format = numbers.FORMAT_NUMBER
		sheet.cell(row=11, column=col + 1, value=row_11).number_format = numbers.FORMAT_NUMBER
		sheet.cell(row=12, column=col + 1, value=row_12).number_format = numbers.FORMAT_NUMBER
		sheet.cell(row=14, column=col + 1, value=row_14).number_format = numbers.FORMAT_DATE_YYYYMMDD2
		
		for row in range(1, 15):
			if row not in [2, 3, 13]:
				sheet.cell(row=row, column=col).value = None
		
		for temp_row in range(15, max_row):
			temp_val = sheet.cell(row=temp_row, column=col).value
			sheet.cell(row=temp_row, column=col + 1, value=temp_val).number_format = numbers.FORMAT_NUMBER
			sheet.cell(row=temp_row, column=col).value = None
	
	sheet.cell(row=1, column=2, value=yesterdays_date.strftime('%m/%d/%y')).number_format = numbers.FORMAT_DATE_YYYYMMDD2
	sheet.cell(row=2, column=2, value='=IFERROR(B9/B11, "N/A")').number_format = numbers.FORMAT_PERCENTAGE
	sheet.cell(row=3, column=2, value='=IFERROR(B7/B12, "N/A")').number_format = numbers.FORMAT_PERCENTAGE
	sheet.cell(row=4, column=2, value=yesterday['conversion']).number_format = numbers.FORMAT_PERCENTAGE
	sheet.cell(row=5, column=2, value=yesterday['bounce']).number_format = numbers.FORMAT_PERCENTAGE
	sheet.cell(row=6, column=2, value=yesterday['engagement']).number_format = numbers.FORMAT_NUMBER
	sheet.cell(row=7, column=2, value=yesterday['visits']).number_format = numbers.FORMAT_NUMBER
	sheet.cell(row=8, column=2, value=yesterday['cta_clicks']).number_format = numbers.FORMAT_NUMBER
	sheet.cell(row=9, column=2, value=yesterday['visitors']).number_format = numbers.FORMAT_NUMBER
	sheet.cell(row=10, column=2, value=yesterday['engagement']).number_format = numbers.FORMAT_NUMBER
	sheet.cell(row=12, column=2, value='=C12').number_format = numbers.FORMAT_NUMBER
	sheet.cell(row=14, column=2, value=yesterdays_date.strftime('%m/%d/%y')).number_format = numbers.FORMAT_DATE_YYYYMMDD2
	
	for country in yesterday['countryCount']:
		if country['country'] in country_idx_manager:
			sheet.cell(row=country_idx_manager[country['country']], column=2, value=country['count']).number_format = numbers.FORMAT_NUMBER
		else:
			sheet.cell(row=(15 + len(country_idx_manager)), column=2, value=country['count']).number_format = numbers.FORMAT_NUMBER
			sheet.cell(row=(15 + len(country_idx_manager)), column=1, value=country['country']).number_format = numbers.FORMAT_GENERAL
			country_idx_manager[country['country']] = country['count']

def print_all_time_to_file(wb: Workbook, all_time: dict[str, Any]):
	sheet = wb['All Time Metrics']
	
	sheet.cell(row=4, column=2, value=all_time['conversion']).number_format = numbers.FORMAT_PERCENTAGE
	sheet.cell(row=5, column=2, value=all_time['bounce']).number_format = numbers.FORMAT_PERCENTAGE
	sheet.cell(row=6, column=2, value=all_time['engagement'])
	sheet.cell(row=7, column=2, value=all_time['visits'])
	sheet.cell(row=8, column=2, value=all_time['cta_clicks'])
	sheet.cell(row=9, column=2, value=all_time['visitors'])
	sheet.cell(row=10, column=2, value=all_time['engagements'])
	
	for row in range(4, 11):
		sheet.cell(row=row, column=2).alignment = Alignment(horizontal='left')
	
	country_idx_manager = {}
	row = 2
	while sheet.cell(row=row, column=4).value:
		country_idx_manager[sheet.cell(row=row, column=4).value] = row
		row += 1
	
	for country in all_time['countryCount']:
		if country['country'] in country_idx_manager:
			sheet.cell(row=country_idx_manager[country['country']], column=5, value=country['count']).number_format = numbers.FORMAT_NUMBER
		else:
			sheet.cell(row=(2 + len(country_idx_manager)), column=4, value=country['country']).number_format = numbers.FORMAT_NUMBER
			sheet.cell(row=(2 + len(country_idx_manager)), column=5, value=country['count']).number_format = numbers.FORMAT_NUMBER
			country_idx_manager[country['country']] = row + len(country_idx_manager)

def print_to_file(yesterday: dict[str, Any], all_time: dict[str, Any], file_path: str):
	wb = load_workbook(file_path)
	
	print_yesterday_to_file(wb, yesterday)
	print_all_time_to_file(wb, all_time)
	
	wb.save(file_path)
